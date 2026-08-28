import io
import json
import base64
import asyncio
import unicodedata
from datetime import datetime
from pathlib import Path

import httpx
import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, Query
from fastapi.responses import Response
from pydantic import BaseModel

from app.core.dependencies import get_current_user
from app.infrastructure.models.usuario import Usuario

router = APIRouter(prefix="/tarifax", tags=["TarifaX"])

DATA_DIR = Path(__file__).parents[4] / "data"
DF1_PATH = DATA_DIR / "TARIFARIO_SICETAC.xlsx"
TEMPLATE_PATH = DATA_DIR / "plantilla_cotizacion_tarifax.xlsx"
# Mapeo categorias internas de la empresa -> tipologia de vehiculo SICETAC.
MAPEO_PATH = DATA_DIR / "tarifax_mapeo_vehiculos.json"

COL_PRECIO_ACTUAL = "TARIFA_CLIENTE"      # precio del cliente (archivo DF2)
COL_PRECIO_SICETAC = "COSTO_TOTAL_VIAJE"  # costo de referencia SICETAC (base DF1)

# Llaves de cruce: (columna en el archivo del cliente DF2, columna en la base SICETAC DF1).
# Definen la RUTA y la CATEGORIA DE VEHICULO que se esta consultando, de modo que el
# merge traiga unicamente la tarifa de esa combinacion (y no todas las de un mismo origen).
JOIN_KEYS = [
    ("ORIGEN", "ORIGEN"),
    ("DESTINO", "DESTINO"),
    ("TIPO_VEHICULO", "TIPO_VEHICULO"),
    ("CARROCERIA", "TIPO_CARROCERIA"),
]
# Sin estas columnas no tiene sentido el cruce por categoria.
REQUIRED_DF2_KEYS = ["ORIGEN", "DESTINO", "TIPO_VEHICULO"]

# Columnas de SICETAC que se arrastran al resultado (ademas del costo).
DF1_EXTRA_COLS = ["DPTO_ORIGEN", "DPTO_DESTINO", "DISTANCIA", "CATEGORIA_VEHICULO"]

COL_DISTANCIA = "DISTANCIA"
# Nombres posibles de la columna de CPK (costo por km) en el tarifario SICETAC.
# El usuario agrega esta columna; si no existe, el CPK se deriva de COSTO/DISTANCIA.
CPK_COL_CANDIDATES = [
    "CPK", "COSTO_KM", "COSTO_POR_KM", "COSTO_POR_KILOMETRO",
    "COSTO_KILOMETRO", "CPK_ORIGEN", "COSTO_X_KM",
]

_df1_cache: pd.DataFrame | None = None
_grouped_cache: dict[tuple, pd.DataFrame] = {}


def _load_df1() -> pd.DataFrame:
    global _df1_cache
    if _df1_cache is None:
        if not DF1_PATH.exists():
            raise HTTPException(status_code=503, detail=f"Archivo base interno no encontrado: {DF1_PATH.name}")
        df = pd.read_excel(DF1_PATH)
        df.columns = [str(c).strip() for c in df.columns]
        _df1_cache = df
    return _df1_cache


def _norm(s: pd.Series) -> pd.Series:
    """Normaliza una columna llave: texto, sin espacios extremos/dobles, MAYUSCULAS y sin acentos.

    Evita que diferencias de mayusculas, espacios o tildes (BOGOTA vs Bogotá) rompan el cruce.
    """
    out = s.astype("string").fillna("").str.strip().str.upper()
    out = out.str.replace(r"\s+", " ", regex=True)
    out = out.map(
        lambda v: unicodedata.normalize("NFKD", v).encode("ascii", "ignore").decode("ascii")
        if isinstance(v, str) else v
    )
    return out


def _grouped_df1(keys: list[tuple[str, str]]) -> pd.DataFrame:
    """Devuelve la base SICETAC colapsada a UNA fila por combinacion de llaves.

    Agrega el costo (promedio) por combinacion — asi el cruce es 1:1 y no explota en
    multiples filas por origen. Cacheado por la firma de columnas de cruce.
    """
    sig = tuple(b for _, b in keys)
    cached = _grouped_cache.get(sig)
    if cached is not None:
        return cached

    df1 = _load_df1().copy()
    norm_cols: list[str] = []
    for i, (_, b) in enumerate(keys):
        nc = f"__k{i}"
        df1[nc] = _norm(df1[b])
        norm_cols.append(nc)

    base = df1[df1[norm_cols].ne("").all(axis=1)]  # descartar llaves vacias

    agg: dict[str, str] = {}
    if COL_PRECIO_SICETAC in df1.columns:
        agg[COL_PRECIO_SICETAC] = "mean"
    for c in DF1_EXTRA_COLS:
        if c in df1.columns:
            agg[c] = "first"

    if agg:
        grouped = base.groupby(norm_cols, as_index=False).agg(agg)
    else:
        grouped = base.groupby(norm_cols, as_index=False).size().drop(columns="size")

    counts = (
        base.groupby(norm_cols, as_index=False)
        .size()
        .rename(columns={"size": "coincidencias_sicetac"})
    )
    grouped = grouped.merge(counts, on=norm_cols, how="left")

    if COL_PRECIO_SICETAC in grouped.columns:
        grouped[COL_PRECIO_SICETAC] = grouped[COL_PRECIO_SICETAC].round(0)

    _grouped_cache[sig] = grouped
    return grouped


def _preview(df: pd.DataFrame, n: int = 25) -> dict:
    """Muestra JSON-safe de las primeras filas para previsualizar en la UI."""
    return {
        "columns": [str(c) for c in df.columns],
        "rows": json.loads(df.head(n).to_json(orient="records", date_format="iso")),
        "total": int(len(df)),
    }


def _cpk_column(df1: pd.DataFrame) -> str | None:
    for c in CPK_COL_CANDIDATES:
        if c in df1.columns:
            return c
    return None


_CPK_PIVOT_COLS = ["MUNICIPIO_ORIGEN", "TIPO_VEHICULO", "CPK_PROMEDIO", "RUTAS_SICETAC"]


def _cpk_por_origen() -> tuple[dict[tuple[str, str], float], pd.DataFrame]:
    """CPK (costo por km) promedio por municipio de ORIGEN Y tipologia de vehiculo.

    El CPK no es igual entre tipologias (tractocamion vs sencillo vs turbo…), por eso
    se clasifica por (ORIGEN, TIPO_VEHICULO). Usa la columna de CPK del tarifario si
    existe; si no, lo deriva de COSTO_TOTAL_VIAJE / DISTANCIA. Devuelve
    (mapa (origen_norm, tipo_veh_norm) -> cpk_promedio, tabla-resumen pivot).
    """
    df1 = _load_df1().copy()
    if "ORIGEN" not in df1.columns or "TIPO_VEHICULO" not in df1.columns:
        return {}, pd.DataFrame(columns=_CPK_PIVOT_COLS)

    df1["__origen"] = _norm(df1["ORIGEN"])
    df1["__tveh"] = _norm(df1["TIPO_VEHICULO"])
    cpk_col = _cpk_column(df1)
    if cpk_col:
        df1["__cpk"] = pd.to_numeric(df1[cpk_col], errors="coerce")
    elif COL_PRECIO_SICETAC in df1.columns and COL_DISTANCIA in df1.columns:
        dist = pd.to_numeric(df1[COL_DISTANCIA], errors="coerce").replace(0, pd.NA)
        df1["__cpk"] = pd.to_numeric(df1[COL_PRECIO_SICETAC], errors="coerce") / dist
    else:
        df1["__cpk"] = pd.NA

    valid = df1[df1["__origen"].ne("") & df1["__tveh"].ne("") & df1["__cpk"].notna()]
    if valid.empty:
        return {}, pd.DataFrame(columns=_CPK_PIVOT_COLS)

    pivot = valid.groupby(["__origen", "__tveh"], as_index=False).agg(
        cpk_promedio=("__cpk", "mean"), rutas=("__cpk", "size")
    )
    nombres_o = df1.groupby("__origen")["ORIGEN"].first()
    nombres_v = df1.groupby("__tveh")["TIPO_VEHICULO"].first()
    pivot["MUNICIPIO_ORIGEN"] = pivot["__origen"].map(nombres_o)
    pivot["TIPO_VEHICULO"] = pivot["__tveh"].map(nombres_v)
    pivot["cpk_promedio"] = pivot["cpk_promedio"].round(2)

    cpk_map = {(r["__origen"], r["__tveh"]): r["cpk_promedio"] for _, r in pivot.iterrows()}
    pivot_out = (
        pivot[["MUNICIPIO_ORIGEN", "TIPO_VEHICULO", "cpk_promedio", "rutas"]]
        .rename(columns={"cpk_promedio": "CPK_PROMEDIO", "rutas": "RUTAS_SICETAC"})
        .sort_values(["MUNICIPIO_ORIGEN", "TIPO_VEHICULO"])
        .reset_index(drop=True)
    )
    return cpk_map, pivot_out


def _distancia_origen_destino() -> dict[tuple[str, str], float]:
    """Distancia promedio por ruta (ORIGEN, DESTINO) desde SICETAC, para estimar
    la tarifa teorica de las rutas que no cruzaron por categoria de vehiculo."""
    df1 = _load_df1().copy()
    if COL_DISTANCIA not in df1.columns or "ORIGEN" not in df1.columns or "DESTINO" not in df1.columns:
        return {}
    df1["__o"] = _norm(df1["ORIGEN"])
    df1["__d"] = _norm(df1["DESTINO"])
    df1["__dist"] = pd.to_numeric(df1[COL_DISTANCIA], errors="coerce")
    valid = df1[df1["__o"].ne("") & df1["__d"].ne("") & df1["__dist"].notna()]
    if valid.empty:
        return {}
    g = valid.groupby(["__o", "__d"], as_index=False)["__dist"].mean()
    return {(r["__o"], r["__d"]): float(r["__dist"]) for _, r in g.iterrows()}


def _load_mapeo_raw() -> dict[str, str]:
    """Mapeo crudo {categoria_interna_empresa: tipo_vehiculo_sicetac}."""
    if MAPEO_PATH.exists():
        try:
            data = json.loads(MAPEO_PATH.read_text(encoding="utf-8"))
            return {str(k): str(v) for k, v in data.items()} if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def _mapeo_norm() -> dict[str, str]:
    """Mapeo con la clave normalizada (para cruzar con el TIPO_VEHICULO del cliente)."""
    out: dict[str, str] = {}
    for interna, sicetac in _load_mapeo_raw().items():
        k = _norm(pd.Series([interna])).iloc[0]
        if k and sicetac:
            out[k] = sicetac
    return out


@router.get("/tipos-sicetac")
async def tipos_sicetac(current_user: Usuario = Depends(get_current_user)):
    """Tipologias de vehiculo tal como SICETAC las nombra (para el mapeo)."""
    df1 = _load_df1()
    if "TIPO_VEHICULO" not in df1.columns:
        return []
    vals = df1["TIPO_VEHICULO"].dropna().astype(str).str.strip()
    return sorted(v for v in vals.unique().tolist() if v)


@router.get("/mapeo-vehiculos")
async def get_mapeo_vehiculos(current_user: Usuario = Depends(get_current_user)):
    return _load_mapeo_raw()


class MapeoVehiculosReq(BaseModel):
    mapeo: dict[str, str]


@router.put("/mapeo-vehiculos")
async def put_mapeo_vehiculos(
    data: MapeoVehiculosReq,
    current_user: Usuario = Depends(get_current_user),
):
    limpio = {str(k).strip(): str(v).strip() for k, v in data.mapeo.items() if str(k).strip() and str(v).strip()}
    MAPEO_PATH.write_text(json.dumps(limpio, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "categorias": len(limpio)}


# ─────────────────────────────────────────────────────────────────────────────
# Cotizacion de almacenamiento: plataformas + inputs configurables -> $/posicion
# ─────────────────────────────────────────────────────────────────────────────
PLATAFORMAS_PATH = DATA_DIR / "tarifax_plataformas.json"
COTIZACION_CONFIG_PATH = DATA_DIR / "tarifax_cotizacion_config.json"

# Distribucion de areas por defecto de una plataforma (m2). La suma son los m2
# utilizados que determinan el % de ocupacion y el prorrateo de servicios.
DEFAULT_AREAS = [
    {"area": "Almacenamiento", "m2": 5200},
    {"area": "Transito (alistamiento - Cross)", "m2": 0},
    {"area": "Muelle (recibo - despacho)", "m2": 0},
    {"area": "Patio de maniobras", "m2": 0},
    {"area": "Devoluciones", "m2": 0},
    {"area": "Despacho ruta nacional", "m2": 0},
    {"area": "Area maquila", "m2": 0},
]

# Productividad por defecto: dimensiona el personal a partir del volumen semanal
# y el rendimiento (unidades por hora-hombre). horas_mes = vol/rend * semanas;
# personas = horas_mes / hr_mensuales.
DEFAULT_PRODUCTIVIDAD = [
    {"actividad": "Recepcion", "volumen_semanal": 0, "unidad": "cajas", "rendimiento_hh": 390},
    {"actividad": "Almacenamiento", "volumen_semanal": 0, "unidad": "estibas", "rendimiento_hh": 12},
    {"actividad": "Alistamiento", "volumen_semanal": 0, "unidad": "cajas", "rendimiento_hh": 0},
    {"actividad": "Despacho", "volumen_semanal": 0, "unidad": "estibas", "rendimiento_hh": 0},
]

# INPUTS pre-configurados (globales). Los valores son un punto de partida
# editable; la ESTRUCTURA de rubros es la que define el costeo por posicion.
DEFAULT_COTIZACION_CONFIG = {
    "parametros": {
        "pallet_largo_m": 1.2,
        "pallet_ancho_m": 1.0,
        "margen_utilidad_pct": 20,
        "ipc_pct": 5.35,
        "smlv": 1750905,
        "aux_transporte": 249095,
        "hr_mensuales": 220,
        "semanas_mes": 4.33,
    },
    "nomina": {"cargos": [
        {"cargo": "Auxiliares de bodega", "cantidad": 0.045, "salario": 2000000, "dotacion": 143000, "carga_prestacional": 840000},
        {"cargo": "Asistente de calidad", "cantidad": 0, "salario": 2000000, "dotacion": 143000, "carga_prestacional": 840000},
        {"cargo": "Montacarguista", "cantidad": 4, "salario": 2089296, "dotacion": 143000, "carga_prestacional": 877504},
        {"cargo": "Asistente de inventarios", "cantidad": 1, "salario": 2451991, "dotacion": 143000, "carga_prestacional": 1029836},
        {"cargo": "Aux de inventarios", "cantidad": 0.045, "salario": 2000000, "dotacion": 143000, "carga_prestacional": 840000},
        {"cargo": "Jefe de bodega", "cantidad": 0.045, "salario": 2338377, "dotacion": 143000, "carga_prestacional": 982118},
        {"cargo": "Vigilante", "cantidad": 0.045, "salario": 2000000, "dotacion": 143000, "carga_prestacional": 840000},
    ]},
    "servicios_publicos": {"items": [
        {"servicio": "Acueducto y alcantarillado", "gasto_total": 2988080},
        {"servicio": "Energia", "gasto_total": 20476396},
        {"servicio": "Gas", "gasto_total": 0},
        {"servicio": "Internet", "gasto_total": 4703111},
    ]},
    "maquinaria": {"incremento_pct": 0, "items": [
        {"item": "Estanteria", "cantidad": 0, "valor": 0},
        {"item": "Montacargas (incluye bateria y cargador)", "cantidad": 0.23, "valor": 5500000},
        {"item": "Bateria montacargas (respaldo)", "cantidad": 0, "valor": 1000000},
        {"item": "Estibadores Manuales", "cantidad": 1, "valor": 150000},
        {"item": "Strech", "cantidad": 0, "valor": 34510},
        {"item": "Estibas", "cantidad": 3000, "valor": 50},
    ]},
    "equipos_tecnologicos": {"incremento_pct": 6, "items": [
        {"item": "Computadores", "cantidad": 0.1, "valor": 125000},
        {"item": "Impresoras regular", "cantidad": 0.1, "valor": 125000},
        {"item": "Impresoras Zebra", "cantidad": 0.1, "valor": 75000},
        {"item": "Scanner", "cantidad": 0, "valor": 125000},
        {"item": "radiofrecuencias", "cantidad": 0.1, "valor": 250000},
        {"item": "Baterias RF", "cantidad": 0.1, "valor": 100000},
        {"item": "Telefonia", "cantidad": 0, "valor": 26000},
    ]},
}


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _load_plataformas() -> list:
    if PLATAFORMAS_PATH.exists():
        try:
            data = json.loads(PLATAFORMAS_PATH.read_text(encoding="utf-8"))
            return data if isinstance(data, list) else []
        except Exception:
            return []
    return []


def _save_plataformas(items: list) -> None:
    PLATAFORMAS_PATH.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")


def _merge_config(cfg: dict) -> dict:
    """Completa un config guardado con las claves nuevas del default (p.ej. parametros)."""
    out = json.loads(json.dumps(DEFAULT_COTIZACION_CONFIG))
    if isinstance(cfg, dict):
        for k, v in cfg.items():
            if k == "parametros" and isinstance(v, dict):
                out["parametros"].update(v)
            else:
                out[k] = v
    return out


def _load_cotizacion_config() -> dict:
    if COTIZACION_CONFIG_PATH.exists():
        try:
            data = json.loads(COTIZACION_CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return _merge_config(data)
        except Exception:
            pass
    return json.loads(json.dumps(DEFAULT_COTIZACION_CONFIG))


def _calcular_rubros(plataforma: dict, tmpl: dict) -> dict:
    """Calcula los rubros (nomina, arriendo, servicios, maquinaria, equipos) y el
    total de operacion mensual de una plataforma. COMPARTIDO por todos los
    servicios de costeo (almacenamiento, cross docking, POP, logistica inversa).

    Los rubros y parametros son PROPIOS de la plataforma; si faltan se usa la
    plantilla como respaldo.
    """
    par = plataforma.get("parametros") or tmpl.get("parametros", {})
    cargos = (plataforma.get("nomina") or tmpl.get("nomina") or {}).get("cargos", [])
    serv_src = (plataforma.get("servicios_publicos") or tmpl.get("servicios_publicos") or {}).get("items", [])
    maq_src = plataforma.get("maquinaria") or tmpl.get("maquinaria") or {}
    eq_src = plataforma.get("equipos_tecnologicos") or tmpl.get("equipos_tecnologicos") or {}

    m2_tot = _num(plataforma.get("m2_totales"))
    arriendo_val = _num(plataforma.get("valor_arriendo"))

    nomina_items = []
    total_nomina = 0.0
    for c in cargos:
        base = _num(c.get("salario")) + _num(c.get("dotacion")) + _num(c.get("carga_prestacional"))
        total = _num(c.get("cantidad")) * base
        total_nomina += total
        nomina_items.append({**c, "salario_dotacion_carga": base, "total": total})

    areas = plataforma.get("areas") or []
    m2_util = sum(_num(a.get("m2")) for a in areas)
    # valor/m2: derivado del arriendo, salvo que se fije un valor/m2 manual
    # (p.ej. tarifa proyectada 2026 distinta del arriendo historico).
    vm2_manual = _num(plataforma.get("valor_m2_manual"))
    valor_m2 = vm2_manual if vm2_manual > 0 else (arriendo_val / m2_tot if m2_tot else 0.0)
    pct_util = m2_util / m2_tot if m2_tot else 0.0
    total_arriendo = m2_util * valor_m2
    areas_calc = [{**a, "asignado": _num(a.get("m2")) * valor_m2} for a in areas]

    serv_items = []
    total_serv = 0.0
    for s in serv_src:
        asignado = _num(s.get("gasto_total")) * pct_util
        total_serv += asignado
        serv_items.append({**s, "asignado": asignado})

    inc_maq = _num(maq_src.get("incremento_pct")) / 100
    maq_items = []
    total_maq = 0.0
    for m in maq_src.get("items", []):
        t = _num(m.get("cantidad")) * _num(m.get("valor"))
        ti = t * (1 + inc_maq)
        total_maq += ti
        maq_items.append({**m, "total": t, "total_incremento": ti})

    inc_eq = _num(eq_src.get("incremento_pct")) / 100
    eq_items = []
    total_eq = 0.0
    for e in eq_src.get("items", []):
        t = _num(e.get("cantidad")) * _num(e.get("valor"))
        ti = t * (1 + inc_eq)
        total_eq += ti
        eq_items.append({**e, "total": t, "total_incremento": ti})

    total_operacion = total_nomina + total_arriendo + total_serv + total_maq + total_eq
    margen = _num(par.get("margen_utilidad_pct")) / 100
    pallet_area = _num(par.get("pallet_largo_m")) * _num(par.get("pallet_ancho_m"))

    def part(x):
        return x / total_operacion if total_operacion else 0.0

    return {
        "nomina": {"items": nomina_items, "total": total_nomina, "participacion": part(total_nomina)},
        "arriendo": {"areas": areas_calc, "m2_utilizados": m2_util, "m2_totales": m2_tot, "valor_m2": valor_m2,
                     "pct_utilizado": pct_util, "total": total_arriendo, "participacion": part(total_arriendo)},
        "servicios_publicos": {"items": serv_items, "pct_utilizado": pct_util,
                               "total": total_serv, "participacion": part(total_serv)},
        "maquinaria": {"items": maq_items, "incremento_pct": inc_maq * 100,
                       "total": total_maq, "participacion": part(total_maq)},
        "equipos_tecnologicos": {"items": eq_items, "incremento_pct": inc_eq * 100,
                                 "total": total_eq, "participacion": part(total_eq)},
        "total_operacion": total_operacion, "margen": margen,
        "m2_utilizados": m2_util, "m2_totales": m2_tot, "pct_utilizado": pct_util,
        "valor_m2": valor_m2, "pallet_area_m2": pallet_area,
    }


def _calcular_productividad(plataforma: dict, tmpl: dict) -> dict:
    """Dimensiona el personal segun volumen semanal y rendimiento (u/HH).

    horas_mes = volumen_semanal / rendimiento_hh * semanas_mes
    personas  = horas_mes / hr_mensuales
    """
    par = plataforma.get("parametros") or tmpl.get("parametros", {})
    hr_mens = _num(par.get("hr_mensuales")) or 220.0
    semanas = _num(par.get("semanas_mes")) or 4.33
    actividades = (plataforma.get("productividad") or {}).get("actividades", [])
    items = []
    total_personas = 0.0
    total_horas = 0.0
    for a in actividades:
        vol = _num(a.get("volumen_semanal"))
        rend = _num(a.get("rendimiento_hh"))
        horas = (vol / rend) * semanas if rend else 0.0
        personas = horas / hr_mens if hr_mens else 0.0
        total_horas += horas
        total_personas += personas
        items.append({**a, "horas_mes": horas, "personas_equiv": personas})
    return {"items": items, "total_horas": total_horas, "total_personas": total_personas,
            "hr_mensuales": hr_mens, "semanas_mes": semanas}


def _calcular_cotizacion(plataforma: dict, config: dict | None = None) -> dict:
    """Costeo de ALMACENAMIENTO -> valor y cobro por POSICION (y por caja/kg)."""
    tmpl = config or _load_cotizacion_config()
    rub = _calcular_rubros(plataforma, tmpl)
    posiciones = _num(plataforma.get("capacidad_posiciones"))
    total = rub["total_operacion"]
    margen = rub["margen"]
    valor_posicion = total / posiciones if posiciones else 0.0
    m2_por_posicion = rub["m2_utilizados"] / posiciones if posiciones else 0.0

    # Denominadores alternativos (como en la hoja: posiciones / cajas / kg)
    cajas = _num(plataforma.get("cajas_movilizadas_mes"))
    kg = _num(plataforma.get("kg_movilizados_mes"))
    valor_caja = total / cajas if cajas else 0.0
    valor_kg = total / kg if kg else 0.0

    prod = _calcular_productividad(plataforma, tmpl)

    return {
        "nomina": rub["nomina"], "arriendo": rub["arriendo"], "servicios_publicos": rub["servicios_publicos"],
        "maquinaria": rub["maquinaria"], "equipos_tecnologicos": rub["equipos_tecnologicos"],
        "productividad": prod,
        "resumen": {
            "total_operacion": total,
            "capacidad_posiciones": posiciones,
            "valor_por_posicion": valor_posicion,
            "margen_utilidad_pct": margen * 100,
            "cobro_por_posicion": valor_posicion * (1 + margen),
            "cajas_movilizadas_mes": cajas,
            "valor_por_caja": valor_caja, "cobro_por_caja": valor_caja * (1 + margen),
            "kg_movilizados_mes": kg,
            "valor_por_kg": valor_kg, "cobro_por_kg": valor_kg * (1 + margen),
            "personas_sugeridas": prod["total_personas"],
            "m2_utilizados": rub["m2_utilizados"], "m2_totales": rub["m2_totales"],
            "pct_utilizado": rub["pct_utilizado"], "valor_m2": rub["valor_m2"],
            "pallet_area_m2": rub["pallet_area_m2"], "m2_por_posicion": m2_por_posicion,
        },
    }


class Plataforma(BaseModel):
    id: int | None = None
    nombre: str
    pais: str = "Colombia"
    ciudad: str = ""
    direccion: str = ""
    posicion: str = ""
    m2_totales: float = 0
    valor_arriendo: float = 0
    valor_m2_manual: float = 0
    capacidad_posiciones: float = 0
    cajas_movilizadas_mes: float = 0
    kg_movilizados_mes: float = 0
    areas: list = []
    productividad: dict = {}
    # Rubros e inputs PROPIOS de cada plataforma (se siembran desde la plantilla)
    parametros: dict = {}
    nomina: dict = {}
    servicios_publicos: dict = {}
    maquinaria: dict = {}
    equipos_tecnologicos: dict = {}
    notas: str = ""


def _sembrar_plataforma(rec: dict) -> dict:
    """Rellena areas, productividad, parametros y rubros vacios con la plantilla."""
    tmpl = _load_cotizacion_config()
    if not rec.get("areas"):
        rec["areas"] = json.loads(json.dumps(DEFAULT_AREAS))
    if not rec.get("productividad") or not rec["productividad"].get("actividades"):
        rec["productividad"] = {"actividades": json.loads(json.dumps(DEFAULT_PRODUCTIVIDAD))}
    for k in ("parametros", "nomina", "servicios_publicos", "maquinaria", "equipos_tecnologicos"):
        if not rec.get(k):
            rec[k] = json.loads(json.dumps(tmpl.get(k, {})))
    return rec


class CotizacionReq(BaseModel):
    plataforma: dict
    config: dict | None = None
    cliente: dict | None = None


@router.get("/plataformas")
async def list_plataformas(current_user: Usuario = Depends(get_current_user)):
    return _load_plataformas()


@router.post("/plataformas")
async def create_plataforma(p: Plataforma, current_user: Usuario = Depends(get_current_user)):
    items = _load_plataformas()
    new_id = (max((int(i.get("id", 0)) for i in items), default=0) + 1)
    rec = _sembrar_plataforma(p.model_dump())
    rec["id"] = new_id
    rec["creado_en"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    items.append(rec)
    _save_plataformas(items)
    return rec


@router.put("/plataformas/{pid}")
async def update_plataforma(pid: int, p: Plataforma, current_user: Usuario = Depends(get_current_user)):
    items = _load_plataformas()
    for i, rec in enumerate(items):
        if int(rec.get("id", -1)) == pid:
            upd = _sembrar_plataforma(p.model_dump())
            upd["id"] = pid
            upd["creado_en"] = rec.get("creado_en")
            items[i] = upd
            _save_plataformas(items)
            return upd
    raise HTTPException(404, "Plataforma no encontrada")


@router.delete("/plataformas/{pid}")
async def delete_plataforma(pid: int, current_user: Usuario = Depends(get_current_user)):
    items = _load_plataformas()
    nuevos = [r for r in items if int(r.get("id", -1)) != pid]
    if len(nuevos) == len(items):
        raise HTTPException(404, "Plataforma no encontrada")
    _save_plataformas(nuevos)
    return {"ok": True}


@router.get("/cotizacion-config")
async def get_cotizacion_config(current_user: Usuario = Depends(get_current_user)):
    return _load_cotizacion_config()


@router.put("/cotizacion-config")
async def put_cotizacion_config(config: dict, current_user: Usuario = Depends(get_current_user)):
    COTIZACION_CONFIG_PATH.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True}


@router.post("/cotizacion-config/reset")
async def reset_cotizacion_config(current_user: Usuario = Depends(get_current_user)):
    return json.loads(json.dumps(DEFAULT_COTIZACION_CONFIG))


@router.get("/cotizacion/areas-default")
async def areas_default(current_user: Usuario = Depends(get_current_user)):
    return json.loads(json.dumps(DEFAULT_AREAS))


@router.post("/cotizacion/calcular")
async def calcular_cotizacion(req: CotizacionReq, current_user: Usuario = Depends(get_current_user)):
    config = _merge_config(req.config) if req.config else _load_cotizacion_config()
    return _calcular_cotizacion(req.plataforma, config)


def _money(x) -> str:
    return "$ " + f"{_num(x):,.0f}".replace(",", ".")


@router.post("/cotizacion/exportar")
async def exportar_cotizacion(req: CotizacionReq, current_user: Usuario = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    config = _merge_config(req.config) if req.config else _load_cotizacion_config()
    r = _calcular_cotizacion(req.plataforma, config)
    p = req.plataforma
    rs = r["resumen"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizacion"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    for col in "CDEFG":
        ws.column_dimensions[col].width = 17
    money = '"$" #,##0'
    hdr = PatternFill("solid", fgColor="D9D9D9")
    bold = Font(bold=True)
    row = [1]

    def emit(vals, *, b=False, fills=None, fmts=None):
        for j, v in enumerate(vals):
            cell = ws.cell(row=row[0], column=j + 1, value=v)
            if b:
                cell.font = bold
            if fills and j in fills:
                cell.fill = hdr
            if fmts and j in fmts:
                cell.number_format = fmts[j]
        row[0] += 1

    emit(["COTIZACION DE ALMACENAMIENTO"], b=True)
    emit(["Plataforma", p.get("nombre", "")])
    emit(["Ubicacion", ", ".join([x for x in [p.get("ciudad", ""), p.get("pais", "")] if x])])
    emit(["Direccion", p.get("direccion", "")])
    emit(["M2 totales", _num(p.get("m2_totales")), "Valor arriendo", _num(p.get("valor_arriendo"))], fmts={3: money})
    emit(["Valor / M2", rs["valor_m2"], "Capacidad (posiciones)", rs["capacidad_posiciones"]], fmts={1: money})
    row[0] += 1

    emit(["NOMINA", "CARGO", "CANT", "SALARIO", "DOTACION", "CARGA PREST.", "SUBTOTAL"], b=True, fills={0, 1, 2, 3, 4, 5, 6})
    for it in r["nomina"]["items"]:
        emit(["", it.get("cargo", ""), _num(it.get("cantidad")), _num(it.get("salario")),
              _num(it.get("dotacion")), _num(it.get("carga_prestacional")), it["total"]],
             fmts={3: money, 4: money, 5: money, 6: money})
    emit(["", "TOTAL NOMINA", "", "", "", "", r["nomina"]["total"]], b=True, fmts={6: money})
    row[0] += 1

    emit(["ARRIENDO", "AREA", "M2", "COSTO ASIGNADO", "", "", ""], b=True, fills={0, 1, 2, 3})
    for a in r["arriendo"]["areas"]:
        emit(["", a.get("area", ""), _num(a.get("m2")), a.get("asignado", 0)], fmts={3: money})
    emit(["", f"M2 utilizados ({r['arriendo']['pct_utilizado']*100:.1f}% de {r['arriendo']['m2_totales']:.0f})",
          r["arriendo"]["m2_utilizados"], "", "Valor/M2", rs["valor_m2"], r["arriendo"]["total"]],
         b=True, fmts={5: money, 6: money})
    row[0] += 1

    emit(["SERVICIOS", "SERVICIO", "GASTO TOTAL", "% M2 UTIL", "ASIGNADO", "", ""], b=True, fills={0, 1, 2, 3, 4})
    for s in r["servicios_publicos"]["items"]:
        emit(["", s.get("servicio", ""), _num(s.get("gasto_total")),
              f"{r['servicios_publicos']['pct_utilizado']*100:.1f}%", s["asignado"]], fmts={2: money, 4: money})
    emit(["", "TOTAL SERVICIOS", "", "", r["servicios_publicos"]["total"]], b=True, fmts={4: money})
    row[0] += 1

    emit([f"MAQUINARIA (+{r['maquinaria']['incremento_pct']:.0f}%)", "ITEM", "CANT", "VALOR", "TOTAL", "TOTAL+INCR", ""], b=True, fills={0, 1, 2, 3, 4, 5})
    for m in r["maquinaria"]["items"]:
        emit(["", m.get("item", ""), _num(m.get("cantidad")), _num(m.get("valor")), m["total"], m["total_incremento"]], fmts={3: money, 4: money, 5: money})
    emit(["", "TOTAL MAQUINARIA", "", "", "", r["maquinaria"]["total"]], b=True, fmts={5: money})
    row[0] += 1

    emit([f"EQUIPOS TEC. (+{r['equipos_tecnologicos']['incremento_pct']:.0f}%)", "ITEM", "CANT", "VALOR", "TOTAL", "TOTAL+INCR", ""], b=True, fills={0, 1, 2, 3, 4, 5})
    for e in r["equipos_tecnologicos"]["items"]:
        emit(["", e.get("item", ""), _num(e.get("cantidad")), _num(e.get("valor")), e["total"], e["total_incremento"]], fmts={3: money, 4: money, 5: money})
    emit(["", "TOTAL EQUIPOS", "", "", "", r["equipos_tecnologicos"]["total"]], b=True, fmts={5: money})
    row[0] += 2

    emit(["RESUMEN"], b=True, fills={0})
    emit(["", "Nomina", r["nomina"]["total"], f"{r['nomina']['participacion']*100:.1f}%"], fmts={2: money})
    emit(["", "Arriendo", r["arriendo"]["total"], f"{r['arriendo']['participacion']*100:.1f}%"], fmts={2: money})
    emit(["", "Servicios publicos", r["servicios_publicos"]["total"], f"{r['servicios_publicos']['participacion']*100:.1f}%"], fmts={2: money})
    emit(["", "Maquinaria y equipo", r["maquinaria"]["total"], f"{r['maquinaria']['participacion']*100:.1f}%"], fmts={2: money})
    emit(["", "Equipos tecnologicos", r["equipos_tecnologicos"]["total"], f"{r['equipos_tecnologicos']['participacion']*100:.1f}%"], fmts={2: money})
    emit(["", "TOTAL OPERACION MENSUAL", rs["total_operacion"]], b=True, fmts={2: money})
    row[0] += 1

    # Valor por unidad (posiciones / cajas / kg movilizados)
    emit(["UNIDAD", "DENOMINADOR", "CANTIDAD/MES", "COSTO UNITARIO", f"COBRO (+{rs['margen_utilidad_pct']:.0f}%)"], b=True, fills={0, 1, 2, 3, 4})
    emit(["", "Posicion de almacenamiento", rs["capacidad_posiciones"], rs["valor_por_posicion"], rs["cobro_por_posicion"]], fmts={3: money, 4: money})
    if rs.get("cajas_movilizadas_mes"):
        emit(["", "Caja movilizada", rs["cajas_movilizadas_mes"], rs["valor_por_caja"], rs["cobro_por_caja"]], fmts={3: money, 4: money})
    if rs.get("kg_movilizados_mes"):
        emit(["", "Kilo movilizado", rs["kg_movilizados_mes"], rs["valor_por_kg"], rs["cobro_por_kg"]], fmts={3: money, 4: money})
    row[0] += 1

    # Productividad (dimensionamiento de personal)
    prod = r.get("productividad", {})
    if prod.get("items"):
        emit(["PRODUCTIVIDAD", "ACTIVIDAD", "VOL/SEMANA", "REND (u/HH)", "HORAS/MES", "PERSONAS"], b=True, fills={0, 1, 2, 3, 4, 5})
        for a in prod["items"]:
            emit(["", a.get("actividad", ""), _num(a.get("volumen_semanal")), _num(a.get("rendimiento_hh")),
                  round(a.get("horas_mes", 0), 1), round(a.get("personas_equiv", 0), 2)])
        emit(["", "TOTAL PERSONAS SUGERIDAS", "", "", round(prod.get("total_horas", 0), 1), round(prod.get("total_personas", 0), 2)], b=True)

    for rr in ws.iter_rows():
        for c in rr:
            if isinstance(c.value, str) and c.column == 2:
                c.alignment = Alignment(wrap_text=False)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre = (p.get("nombre") or "plataforma").replace(" ", "_")
    filename = f"Cotizacion_{nombre}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return {
        "filename": filename,
        "file_base64": base64.b64encode(output.read()).decode("utf-8"),
        "resumen": rs,
    }


@router.post("/cotizacion/pdf")
async def pdf_cotizacion(req: CotizacionReq, current_user: Usuario = Depends(get_current_user)):
    """Genera una cotizacion FORMAL (PDF) para el cliente: oferta por posicion,
    que incluye el servicio y un desglose de la estructura de costos."""
    from fpdf import FPDF

    config = _merge_config(req.config) if req.config else _load_cotizacion_config()
    r = _calcular_cotizacion(req.plataforma, config)
    p = req.plataforma
    rs = r["resumen"]
    cli = req.cliente or {}

    GREEN = (54, 158, 77)
    DARK = (31, 97, 48)
    GREY = (100, 116, 139)
    LIGHT = (238, 244, 240)

    def lat(s):
        return str(s).encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF(format="A4", unit="mm")
    pdf.set_auto_page_break(True, margin=16)
    pdf.add_page()
    W = 210
    M = 14

    # Banda superior
    pdf.set_fill_color(*GREEN)
    pdf.rect(0, 0, W, 30, "F")
    pdf.set_xy(M, 8)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 20)
    pdf.cell(120, 9, lat("COTIZACION"), ln=1)
    pdf.set_x(M)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(120, 5, lat("Servicio de almacenamiento y operacion logistica"), ln=1)
    # Numero / fecha a la derecha
    hoy = datetime.now()
    num = f"COT-{hoy.strftime('%Y%m%d-%H%M')}"
    pdf.set_xy(W - 78, 9)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(64, 6, lat(f"No. {num}"), align="R", ln=2)
    pdf.set_x(W - 78)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(64, 5, lat(f"Fecha: {hoy.strftime('%Y-%m-%d')}"), align="R", ln=2)
    pdf.set_x(W - 78)
    pdf.cell(64, 5, lat("Validez: 30 dias"), align="R", ln=2)

    pdf.set_text_color(30, 41, 59)
    pdf.set_y(37)

    def seccion(titulo):
        pdf.ln(2)
        pdf.set_font("helvetica", "B", 11)
        pdf.set_text_color(*DARK)
        pdf.set_x(M)
        pdf.cell(0, 7, lat(titulo), ln=1)
        pdf.set_draw_color(*GREEN)
        pdf.set_line_width(0.5)
        pdf.line(M, pdf.get_y(), W - M, pdf.get_y())
        pdf.ln(1.5)
        pdf.set_text_color(30, 41, 59)

    def kv(label, value, x, w_label=38, w_val=52):
        pdf.set_x(x)
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(*GREY)
        pdf.cell(w_label, 6, lat(label), ln=0)
        pdf.set_font("helvetica", "", 9)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(w_val, 6, lat(value), ln=0)

    # Datos cliente / operacion
    seccion("Cliente y operacion")
    y0 = pdf.get_y()
    kv("Cliente:", cli.get("nombre", "________________________"), M)
    kv("Plataforma:", p.get("nombre", ""), W / 2)
    pdf.ln(6)
    kv("Contacto:", cli.get("contacto", ""), M)
    ubic = ", ".join([x for x in [p.get("ciudad", ""), p.get("pais", "")] if x])
    kv("Ubicacion:", ubic, W / 2)
    pdf.ln(6)
    kv("NIT / ID:", cli.get("nit", ""), M)
    kv("Direccion:", p.get("direccion", ""), W / 2)
    pdf.ln(8)

    # Caja de oferta destacada
    seccion("Nuestra oferta")
    box_y = pdf.get_y()
    pdf.set_fill_color(*LIGHT)
    pdf.rect(M, box_y, W - 2 * M, 26, "F")
    pdf.set_xy(M + 4, box_y + 4)
    pdf.set_font("helvetica", "B", 9)
    pdf.set_text_color(*GREY)
    pdf.cell(90, 5, lat("TARIFA POR POSICION / MES"), ln=2)
    pdf.set_x(M + 4)
    pdf.set_font("helvetica", "B", 26)
    pdf.set_text_color(*DARK)
    pdf.cell(100, 12, lat(_money(rs["cobro_por_posicion"])), ln=2)
    # Datos a la derecha de la caja
    pdf.set_xy(W / 2 + 6, box_y + 4)
    pdf.set_font("helvetica", "", 9)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(80, 5, lat(f"Posiciones cotizadas: {rs['capacidad_posiciones']:,.0f}".replace(",", ".")), ln=2)
    pdf.set_x(W / 2 + 6)
    valor_mes = rs["cobro_por_posicion"] * rs["capacidad_posiciones"]
    pdf.cell(80, 5, lat(f"Valor mensual estimado: {_money(valor_mes)}"), ln=2)
    pdf.set_x(W / 2 + 6)
    pdf.cell(80, 5, lat(f"Dimension posicion (pallet): {_num(config['parametros'].get('pallet_largo_m')):.2g} x {_num(config['parametros'].get('pallet_ancho_m')):.2g} m"), ln=2)
    pdf.set_y(box_y + 30)

    # El servicio incluye
    seccion("El servicio incluye")
    incluye = [
        ("Talento humano de bodega", "Auxiliares, montacarguistas, inventarios, calidad y supervision."),
        ("Espacio de almacenamiento", f"{rs['m2_utilizados']:,.0f} m2 en operacion ({rs['pct_utilizado']*100:.1f}% de la bodega).".replace(",", ".")),
        ("Servicios publicos", "Energia, acueducto, gas e internet prorrateados a la operacion."),
        ("Maquinaria y equipo", "Montacargas, estibadores y estibas para el manejo de carga."),
        ("Equipos tecnologicos", "Computadores, impresoras, radiofrecuencias y sistema WMS."),
    ]
    for t, d in incluye:
        pdf.set_x(M)
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(*DARK)
        pdf.cell(3, 5, lat("-"), ln=0)
        pdf.cell(52, 5, lat(t), ln=0)
        pdf.set_font("helvetica", "", 9)
        pdf.set_text_color(30, 41, 59)
        pdf.multi_cell(W - 2 * M - 55, 5, lat(d))
    pdf.ln(2)

    # Estructura de costos (resumen por concepto)
    seccion("Estructura de la operacion (mensual)")
    filas = [
        ("Nomina", r["nomina"]["total"], r["nomina"]["participacion"]),
        ("Arriendo de espacio", r["arriendo"]["total"], r["arriendo"]["participacion"]),
        ("Servicios publicos", r["servicios_publicos"]["total"], r["servicios_publicos"]["participacion"]),
        ("Maquinaria y equipo", r["maquinaria"]["total"], r["maquinaria"]["participacion"]),
        ("Equipos tecnologicos", r["equipos_tecnologicos"]["total"], r["equipos_tecnologicos"]["participacion"]),
    ]
    pdf.set_x(M)
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(*GREEN)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(90, 7, lat("  Concepto"), border=0, fill=True, ln=0)
    pdf.cell(56, 7, lat("Valor mensual"), border=0, fill=True, align="R", ln=0)
    pdf.cell(36, 7, lat("Participacion  "), border=0, fill=True, align="R", ln=1)
    pdf.set_text_color(30, 41, 59)
    alt = False
    for concepto, val, pp in filas:
        pdf.set_x(M)
        pdf.set_fill_color(245, 248, 246) if alt else pdf.set_fill_color(255, 255, 255)
        pdf.set_font("helvetica", "", 9)
        pdf.cell(90, 6.5, lat("  " + concepto), border=0, fill=True, ln=0)
        pdf.cell(56, 6.5, lat(_money(val)), border=0, fill=True, align="R", ln=0)
        pdf.cell(36, 6.5, lat(f"{pp*100:.1f}%  "), border=0, fill=True, align="R", ln=1)
        alt = not alt
    pdf.set_x(M)
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(*LIGHT)
    pdf.cell(90, 7, lat("  TOTAL OPERACION"), border=0, fill=True, ln=0)
    pdf.cell(56, 7, lat(_money(rs["total_operacion"])), border=0, fill=True, align="R", ln=0)
    pdf.cell(36, 7, lat("100%  "), border=0, fill=True, align="R", ln=1)
    pdf.set_x(M)
    pdf.cell(90, 7, lat("  Costo por posicion"), border=0, ln=0)
    pdf.cell(56, 7, lat(_money(rs["valor_por_posicion"])), border=0, align="R", ln=0)
    pdf.cell(36, 7, lat(f"margen {rs['margen_utilidad_pct']:.0f}%  "), border=0, align="R", ln=1)

    # Tarifas alternativas (caja / kg movilizados) cuando aplican
    otras = []
    if rs.get("cajas_movilizadas_mes"):
        otras.append(("Cobro por caja movilizada", rs["cobro_por_caja"]))
    if rs.get("kg_movilizados_mes"):
        otras.append(("Cobro por kilo movilizado", rs["cobro_por_kg"]))
    if otras:
        pdf.ln(2)
        seccion("Tarifas alternativas")
        for lbl, val in otras:
            pdf.set_x(M)
            pdf.set_font("helvetica", "", 9)
            pdf.cell(120, 6, lat("- " + lbl), ln=0)
            pdf.set_font("helvetica", "B", 9)
            pdf.cell(60, 6, lat(_money(val)), align="R", ln=1)

    # Terminos
    seccion("Condiciones comerciales")
    pdf.set_font("helvetica", "", 8.5)
    pdf.set_text_color(*GREY)
    terminos = (
        "Los valores estan expresados en pesos colombianos (COP) y no incluyen IVA. "
        "La tarifa por posicion aplica sobre las posiciones contratadas mensualmente. "
        "Esta cotizacion tiene una validez de 30 dias a partir de la fecha de emision y esta "
        "sujeta a la firma del contrato de prestacion de servicios logisticos. Servicios "
        "adicionales (maquila, cross-docking, transporte) se cotizan por separado."
    )
    pdf.set_x(M)
    pdf.multi_cell(W - 2 * M, 4.5, lat(terminos))

    out = pdf.output()
    data = bytes(out)
    nombre = (p.get("nombre") or "plataforma").replace(" ", "_")
    filename = f"Cotizacion_{nombre}_{hoy.strftime('%Y%m%d_%H%M%S')}.pdf"
    return {
        "filename": filename,
        "file_base64": base64.b64encode(data).decode("utf-8"),
        "resumen": rs,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Costeo GENERICO por servicio: cross docking, POP, logistica inversa.
# Mismo motor de rubros que almacenamiento; la UNIDAD DE COBRO es configurable
# por plataforma (lista de {unidad, cantidad} -> costo y cobro por cada unidad).
# ─────────────────────────────────────────────────────────────────────────────
SERVICIOS_COSTEO = {
    "cross_docking": {
        "label": "Cross Docking",
        "subtitulo": "Servicio de cross docking y operacion logistica",
        "unidades": ["Pallet cruzado", "Caja", "Tonelada", "Kilo", "Movimiento"],
    },
    "pop": {
        "label": "POP",
        "subtitulo": "Servicio de armado y procesamiento de material POP",
        "unidades": ["Unidad procesada"],
    },
    "logistica_inversa": {
        "label": "Logistica Inversa",
        "subtitulo": "Servicio de logistica inversa y gestion de devoluciones",
        "unidades": ["Devolucion", "Unidad", "Caja", "Pallet", "Orden", "Kilo", "Tonelada"],
    },
}


def _svc(servicio: str) -> dict:
    s = SERVICIOS_COSTEO.get(servicio)
    if not s:
        raise HTTPException(404, f"Servicio de costeo no valido: {servicio}")
    return s


def _svc_paths(servicio: str):
    _svc(servicio)
    return (DATA_DIR / f"tarifax_{servicio}_plataformas.json",
            DATA_DIR / f"tarifax_{servicio}_config.json")


def _load_json_list(path) -> list:
    if path.exists():
        try:
            d = json.loads(path.read_text(encoding="utf-8"))
            return d if isinstance(d, list) else []
        except Exception:
            return []
    return []


def _svc_default_config() -> dict:
    # Mismos rubros base que almacenamiento (personal, servicios, maquinaria, equipos).
    return json.loads(json.dumps(DEFAULT_COTIZACION_CONFIG))


def _load_svc_config(servicio: str) -> dict:
    _, cfg_path = _svc_paths(servicio)
    if cfg_path.exists():
        try:
            data = json.loads(cfg_path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return _merge_config(data)
        except Exception:
            pass
    return _svc_default_config()


def _sembrar_svc(rec: dict, servicio: str) -> dict:
    tmpl = _load_svc_config(servicio)
    if not rec.get("areas"):
        rec["areas"] = json.loads(json.dumps(DEFAULT_AREAS))
    if not rec.get("unidades"):
        rec["unidades"] = [{"unidad": u, "cantidad": 0} for u in _svc(servicio)["unidades"]]
    for k in ("parametros", "nomina", "servicios_publicos", "maquinaria", "equipos_tecnologicos"):
        if not rec.get(k):
            rec[k] = json.loads(json.dumps(tmpl.get(k, {})))
    return rec


def _resolver_base(plataforma: dict) -> dict:
    """Si la plataforma de servicio esta VINCULADA a una bodega de almacenamiento,
    hereda de ella los datos de la instalacion (m2, arriendo, valor/m2, ubicacion).
    Los rubros, areas y unidades siguen siendo propios del servicio."""
    bid = plataforma.get("base_almacenamiento_id")
    if not bid:
        return plataforma
    try:
        bases = {int(b.get("id")): b for b in _load_plataformas() if b.get("id") is not None}
        base = bases.get(int(bid))
    except (TypeError, ValueError):
        base = None
    if not base:
        return plataforma
    p = dict(plataforma)
    for k in ("m2_totales", "valor_arriendo", "valor_m2_manual", "pais", "ciudad", "direccion"):
        p[k] = base.get(k, p.get(k))
    p["_base_nombre"] = base.get("nombre", "")
    return p


def _calcular_costeo(plataforma: dict, config: dict | None, servicio: str) -> dict:
    """Costeo generico -> costo y cobro por CADA unidad de cobro configurada."""
    plataforma = _resolver_base(plataforma)
    tmpl = config or _load_svc_config(servicio)
    rub = _calcular_rubros(plataforma, tmpl)
    total = rub["total_operacion"]
    margen = rub["margen"]
    unidades = []
    for u in (plataforma.get("unidades") or []):
        cant = _num(u.get("cantidad"))
        costo = total / cant if cant else 0.0
        unidades.append({"unidad": u.get("unidad", ""), "cantidad": cant,
                         "costo_unitario": costo, "cobro_unitario": costo * (1 + margen)})
    return {
        "nomina": rub["nomina"], "arriendo": rub["arriendo"], "servicios_publicos": rub["servicios_publicos"],
        "maquinaria": rub["maquinaria"], "equipos_tecnologicos": rub["equipos_tecnologicos"],
        "resumen": {
            "servicio": servicio, "total_operacion": total, "margen_utilidad_pct": margen * 100,
            "unidades": unidades, "m2_utilizados": rub["m2_utilizados"], "m2_totales": rub["m2_totales"],
            "pct_utilizado": rub["pct_utilizado"], "valor_m2": rub["valor_m2"],
            "base_almacenamiento_id": plataforma.get("base_almacenamiento_id"),
            "base_nombre": plataforma.get("_base_nombre", ""),
        },
    }


class PlataformaCosteo(BaseModel):
    id: int | None = None
    nombre: str
    base_almacenamiento_id: int | None = None
    pais: str = "Colombia"
    ciudad: str = ""
    direccion: str = ""
    posicion: str = ""
    m2_totales: float = 0
    valor_arriendo: float = 0
    valor_m2_manual: float = 0
    unidades: list = []
    areas: list = []
    parametros: dict = {}
    nomina: dict = {}
    servicios_publicos: dict = {}
    maquinaria: dict = {}
    equipos_tecnologicos: dict = {}
    notas: str = ""


class CosteoReq(BaseModel):
    plataforma: dict
    config: dict | None = None
    cliente: dict | None = None


@router.get("/servicios/{servicio}/meta")
async def svc_meta(servicio: str, current_user: Usuario = Depends(get_current_user)):
    s = _svc(servicio)
    return {"servicio": servicio, "label": s["label"], "subtitulo": s["subtitulo"], "unidades_sugeridas": s["unidades"]}


@router.get("/servicios/{servicio}/plataformas")
async def svc_list(servicio: str, current_user: Usuario = Depends(get_current_user)):
    ppath, _ = _svc_paths(servicio)
    return _load_json_list(ppath)


@router.post("/servicios/{servicio}/plataformas")
async def svc_create(servicio: str, p: PlataformaCosteo, current_user: Usuario = Depends(get_current_user)):
    ppath, _ = _svc_paths(servicio)
    items = _load_json_list(ppath)
    new_id = (max((int(i.get("id", 0)) for i in items), default=0) + 1)
    rec = _sembrar_svc(p.model_dump(), servicio)
    rec["id"] = new_id
    rec["creado_en"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    items.append(rec)
    ppath.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    return rec


@router.put("/servicios/{servicio}/plataformas/{pid}")
async def svc_update(servicio: str, pid: int, p: PlataformaCosteo, current_user: Usuario = Depends(get_current_user)):
    ppath, _ = _svc_paths(servicio)
    items = _load_json_list(ppath)
    for i, rec in enumerate(items):
        if int(rec.get("id", -1)) == pid:
            upd = _sembrar_svc(p.model_dump(), servicio)
            upd["id"] = pid
            upd["creado_en"] = rec.get("creado_en")
            items[i] = upd
            ppath.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
            return upd
    raise HTTPException(404, "Plataforma no encontrada")


@router.delete("/servicios/{servicio}/plataformas/{pid}")
async def svc_delete(servicio: str, pid: int, current_user: Usuario = Depends(get_current_user)):
    ppath, _ = _svc_paths(servicio)
    items = _load_json_list(ppath)
    nuevos = [r for r in items if int(r.get("id", -1)) != pid]
    if len(nuevos) == len(items):
        raise HTTPException(404, "Plataforma no encontrada")
    ppath.write_text(json.dumps(nuevos, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True}


@router.get("/servicios/{servicio}/config")
async def svc_get_config(servicio: str, current_user: Usuario = Depends(get_current_user)):
    return _load_svc_config(servicio)


@router.put("/servicios/{servicio}/config")
async def svc_put_config(servicio: str, config: dict, current_user: Usuario = Depends(get_current_user)):
    _, cfg_path = _svc_paths(servicio)
    cfg_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True}


@router.post("/servicios/{servicio}/config/reset")
async def svc_reset_config(servicio: str, current_user: Usuario = Depends(get_current_user)):
    _svc(servicio)
    return _svc_default_config()


@router.post("/servicios/{servicio}/cotizacion/calcular")
async def svc_calcular(servicio: str, req: CosteoReq, current_user: Usuario = Depends(get_current_user)):
    config = _merge_config(req.config) if req.config else _load_svc_config(servicio)
    return _calcular_costeo(req.plataforma, config, servicio)


@router.post("/servicios/{servicio}/cotizacion/exportar")
async def svc_exportar(servicio: str, req: CosteoReq, current_user: Usuario = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    s = _svc(servicio)
    config = _merge_config(req.config) if req.config else _load_svc_config(servicio)
    r = _calcular_costeo(req.plataforma, config, servicio)
    p = _resolver_base(req.plataforma)
    rs = r["resumen"]

    wb = Workbook()
    ws = wb.active
    ws.title = s["label"][:31]
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    for col in "CDEFG":
        ws.column_dimensions[col].width = 17
    money = '"$" #,##0'
    hdr = PatternFill("solid", fgColor="D9D9D9")
    bold = Font(bold=True)
    row = [1]

    def emit(vals, *, b=False, fills=None, fmts=None):
        for j, v in enumerate(vals):
            cell = ws.cell(row=row[0], column=j + 1, value=v)
            if b:
                cell.font = bold
            if fills and j in fills:
                cell.fill = hdr
            if fmts and j in fmts:
                cell.number_format = fmts[j]
        row[0] += 1

    emit([f"COTIZACION {s['label'].upper()}"], b=True)
    emit(["Plataforma", p.get("nombre", "")])
    emit(["Ubicacion", ", ".join([x for x in [p.get("ciudad", ""), p.get("pais", "")] if x])])
    emit(["Direccion", p.get("direccion", "")])
    row[0] += 1

    emit(["NOMINA", "CARGO", "CANT", "SALARIO", "DOTACION", "CARGA PREST.", "SUBTOTAL"], b=True, fills={0, 1, 2, 3, 4, 5, 6})
    for it in r["nomina"]["items"]:
        emit(["", it.get("cargo", ""), _num(it.get("cantidad")), _num(it.get("salario")),
              _num(it.get("dotacion")), _num(it.get("carga_prestacional")), it["total"]],
             fmts={3: money, 4: money, 5: money, 6: money})
    emit(["", "TOTAL NOMINA", "", "", "", "", r["nomina"]["total"]], b=True, fmts={6: money})
    row[0] += 1

    emit(["ARRIENDO", "AREA", "M2", "COSTO ASIGNADO"], b=True, fills={0, 1, 2, 3})
    for a in r["arriendo"]["areas"]:
        emit(["", a.get("area", ""), _num(a.get("m2")), a.get("asignado", 0)], fmts={3: money})
    emit(["", "TOTAL ARRIENDO", "", r["arriendo"]["total"]], b=True, fmts={3: money})
    row[0] += 1

    emit(["SERVICIOS", "SERVICIO", "GASTO TOTAL", "% M2 UTIL", "ASIGNADO"], b=True, fills={0, 1, 2, 3, 4})
    for sv in r["servicios_publicos"]["items"]:
        emit(["", sv.get("servicio", ""), _num(sv.get("gasto_total")),
              f"{r['servicios_publicos']['pct_utilizado']*100:.1f}%", sv["asignado"]], fmts={2: money, 4: money})
    emit(["", "TOTAL SERVICIOS", "", "", r["servicios_publicos"]["total"]], b=True, fmts={4: money})
    row[0] += 1

    emit([f"MAQUINARIA (+{r['maquinaria']['incremento_pct']:.0f}%)", "ITEM", "CANT", "VALOR", "TOTAL", "TOTAL+INCR"], b=True, fills={0, 1, 2, 3, 4, 5})
    for m in r["maquinaria"]["items"]:
        emit(["", m.get("item", ""), _num(m.get("cantidad")), _num(m.get("valor")), m["total"], m["total_incremento"]], fmts={3: money, 4: money, 5: money})
    emit(["", "TOTAL MAQUINARIA", "", "", "", r["maquinaria"]["total"]], b=True, fmts={5: money})
    row[0] += 1

    emit([f"EQUIPOS TEC. (+{r['equipos_tecnologicos']['incremento_pct']:.0f}%)", "ITEM", "CANT", "VALOR", "TOTAL", "TOTAL+INCR"], b=True, fills={0, 1, 2, 3, 4, 5})
    for e in r["equipos_tecnologicos"]["items"]:
        emit(["", e.get("item", ""), _num(e.get("cantidad")), _num(e.get("valor")), e["total"], e["total_incremento"]], fmts={3: money, 4: money, 5: money})
    emit(["", "TOTAL EQUIPOS", "", "", "", r["equipos_tecnologicos"]["total"]], b=True, fmts={5: money})
    row[0] += 2

    emit(["RESUMEN"], b=True, fills={0})
    for lbl, key in [("Nomina", "nomina"), ("Arriendo", "arriendo"), ("Servicios publicos", "servicios_publicos"),
                     ("Maquinaria y equipo", "maquinaria"), ("Equipos tecnologicos", "equipos_tecnologicos")]:
        emit(["", lbl, r[key]["total"], f"{r[key]['participacion']*100:.1f}%"], fmts={2: money})
    emit(["", "TOTAL OPERACION MENSUAL", rs["total_operacion"]], b=True, fmts={2: money})
    row[0] += 1

    emit(["TARIFAS", "UNIDAD DE COBRO", "CANTIDAD/MES", "COSTO UNITARIO", f"COBRO (+{rs['margen_utilidad_pct']:.0f}%)"], b=True, fills={0, 1, 2, 3, 4})
    for u in rs["unidades"]:
        emit(["", u["unidad"], u["cantidad"], u["costo_unitario"], u["cobro_unitario"]], fmts={3: money, 4: money})

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nombre = (p.get("nombre") or servicio).replace(" ", "_")
    filename = f"Cotizacion_{s['label'].replace(' ', '')}_{nombre}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return {"filename": filename, "file_base64": base64.b64encode(output.read()).decode("utf-8"), "resumen": rs}


@router.post("/servicios/{servicio}/cotizacion/pdf")
async def svc_pdf(servicio: str, req: CosteoReq, current_user: Usuario = Depends(get_current_user)):
    """Cotizacion FORMAL (PDF) para el cliente, con tarifas por unidad de cobro."""
    from fpdf import FPDF

    s = _svc(servicio)
    config = _merge_config(req.config) if req.config else _load_svc_config(servicio)
    r = _calcular_costeo(req.plataforma, config, servicio)
    p = _resolver_base(req.plataforma)
    rs = r["resumen"]
    cli = req.cliente or {}

    GREEN = (54, 158, 77)
    DARK = (31, 97, 48)
    GREY = (100, 116, 139)
    LIGHT = (238, 244, 240)

    def lat(x):
        return str(x).encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF(format="A4", unit="mm")
    pdf.set_auto_page_break(True, margin=16)
    pdf.add_page()
    W, M = 210, 14

    pdf.set_fill_color(*GREEN)
    pdf.rect(0, 0, W, 30, "F")
    pdf.set_xy(M, 8)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", "B", 20)
    pdf.cell(120, 9, lat("COTIZACION"), ln=1)
    pdf.set_x(M)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(120, 5, lat(s["subtitulo"]), ln=1)
    hoy = datetime.now()
    pdf.set_xy(W - 78, 9)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(64, 6, lat(f"No. COT-{hoy.strftime('%Y%m%d-%H%M')}"), align="R", ln=2)
    pdf.set_x(W - 78)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(64, 5, lat(f"Fecha: {hoy.strftime('%Y-%m-%d')}"), align="R", ln=2)
    pdf.set_x(W - 78)
    pdf.cell(64, 5, lat("Validez: 30 dias"), align="R", ln=2)

    pdf.set_text_color(30, 41, 59)
    pdf.set_y(37)

    def seccion(t):
        pdf.ln(2)
        pdf.set_font("helvetica", "B", 11)
        pdf.set_text_color(*DARK)
        pdf.set_x(M)
        pdf.cell(0, 7, lat(t), ln=1)
        pdf.set_draw_color(*GREEN)
        pdf.set_line_width(0.5)
        pdf.line(M, pdf.get_y(), W - M, pdf.get_y())
        pdf.ln(1.5)
        pdf.set_text_color(30, 41, 59)

    def kv(label, value, x, wl=38, wv=52):
        pdf.set_x(x)
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(*GREY)
        pdf.cell(wl, 6, lat(label), ln=0)
        pdf.set_font("helvetica", "", 9)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(wv, 6, lat(value), ln=0)

    seccion("Cliente y operacion")
    kv("Cliente:", cli.get("nombre", "________________________"), M)
    kv("Servicio:", s["label"], W / 2)
    pdf.ln(6)
    kv("Contacto:", cli.get("contacto", ""), M)
    kv("Plataforma:", p.get("nombre", ""), W / 2)
    pdf.ln(6)
    kv("NIT / ID:", cli.get("nit", ""), M)
    kv("Ubicacion:", ", ".join([x for x in [p.get("ciudad", ""), p.get("pais", "")] if x]), W / 2)
    pdf.ln(8)

    # Tarifas por unidad de cobro
    seccion("Tarifas por unidad de cobro")
    pdf.set_x(M)
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(*GREEN)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(78, 7, lat("  Unidad"), fill=True, ln=0)
    pdf.cell(36, 7, lat("Cantidad/mes"), fill=True, align="R", ln=0)
    pdf.cell(68, 7, lat(f"Tarifa unitaria (+{rs['margen_utilidad_pct']:.0f}%)  "), fill=True, align="R", ln=1)
    pdf.set_text_color(30, 41, 59)
    alt = False
    tiene = [u for u in rs["unidades"] if u["cantidad"]]
    for u in (tiene or rs["unidades"]):
        pdf.set_x(M)
        pdf.set_fill_color(245, 248, 246) if alt else pdf.set_fill_color(255, 255, 255)
        pdf.set_font("helvetica", "", 9)
        pdf.cell(78, 6.5, lat("  " + (u["unidad"] or "-")), fill=True, ln=0)
        pdf.cell(36, 6.5, lat(f"{u['cantidad']:,.0f}".replace(",", ".")), fill=True, align="R", ln=0)
        pdf.set_font("helvetica", "B", 9)
        pdf.cell(68, 6.5, lat(_money(u["cobro_unitario"]) + "  "), fill=True, align="R", ln=1)
        alt = not alt
    pdf.ln(2)

    seccion("El servicio incluye")
    incluye = [
        ("Talento humano", "Personal operativo, de inventarios, calidad y supervision."),
        ("Espacio de operacion", f"{rs['m2_utilizados']:,.0f} m2 asignados a la operacion.".replace(",", ".")),
        ("Servicios publicos", "Energia, acueducto, gas e internet prorrateados."),
        ("Maquinaria y equipo", "Montacargas, estibadores y elementos de manejo de carga."),
        ("Equipos tecnologicos", "Computadores, impresoras, radiofrecuencias y sistema WMS."),
    ]
    for t, d in incluye:
        pdf.set_x(M)
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(*DARK)
        pdf.cell(3, 5, lat("-"), ln=0)
        pdf.cell(46, 5, lat(t), ln=0)
        pdf.set_font("helvetica", "", 9)
        pdf.set_text_color(30, 41, 59)
        pdf.multi_cell(W - 2 * M - 49, 5, lat(d))
    pdf.ln(2)

    seccion("Estructura de la operacion (mensual)")
    pdf.set_x(M)
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(*GREEN)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(90, 7, lat("  Concepto"), fill=True, ln=0)
    pdf.cell(56, 7, lat("Valor mensual"), fill=True, align="R", ln=0)
    pdf.cell(36, 7, lat("Participacion  "), fill=True, align="R", ln=1)
    pdf.set_text_color(30, 41, 59)
    alt = False
    for lbl, key in [("Nomina", "nomina"), ("Arriendo de espacio", "arriendo"), ("Servicios publicos", "servicios_publicos"),
                     ("Maquinaria y equipo", "maquinaria"), ("Equipos tecnologicos", "equipos_tecnologicos")]:
        pdf.set_x(M)
        pdf.set_fill_color(245, 248, 246) if alt else pdf.set_fill_color(255, 255, 255)
        pdf.set_font("helvetica", "", 9)
        pdf.cell(90, 6.5, lat("  " + lbl), fill=True, ln=0)
        pdf.cell(56, 6.5, lat(_money(r[key]["total"])), fill=True, align="R", ln=0)
        pdf.cell(36, 6.5, lat(f"{r[key]['participacion']*100:.1f}%  "), fill=True, align="R", ln=1)
        alt = not alt
    pdf.set_x(M)
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(*LIGHT)
    pdf.cell(90, 7, lat("  TOTAL OPERACION"), fill=True, ln=0)
    pdf.cell(56, 7, lat(_money(rs["total_operacion"])), fill=True, align="R", ln=0)
    pdf.cell(36, 7, lat("100%  "), fill=True, align="R", ln=1)

    seccion("Condiciones comerciales")
    pdf.set_font("helvetica", "", 8.5)
    pdf.set_text_color(*GREY)
    pdf.set_x(M)
    pdf.multi_cell(W - 2 * M, 4.5, lat(
        "Los valores estan expresados en pesos colombianos (COP) y no incluyen IVA. Las tarifas por "
        "unidad aplican sobre los volumenes contratados mensualmente. Esta cotizacion tiene una validez "
        "de 30 dias a partir de la fecha de emision y esta sujeta a la firma del contrato de prestacion "
        "de servicios logisticos. Servicios adicionales se cotizan por separado."))

    data = bytes(pdf.output())
    nombre = (p.get("nombre") or servicio).replace(" ", "_")
    filename = f"Cotizacion_{s['label'].replace(' ', '')}_{nombre}_{hoy.strftime('%Y%m%d_%H%M%S')}.pdf"
    return {"filename": filename, "file_base64": base64.b64encode(data).decode("utf-8"), "resumen": rs}


# ─── Ruteo por carretera (geocodificacion Nominatim + ruta OSRM) ──────────────
# OSRM implementa Dijkstra sobre Contraction Hierarchies para calcular la ruta
# minima por la red vial real. Nominatim resuelve nombres de lugar -> lat/lon.
_NOMINATIM = "https://nominatim.openstreetmap.org/search"
_OSRM = "https://router.project-osrm.org/route/v1/driving"
_UA = {"User-Agent": "TarifaX/1.0 (contacto@empresa.com)"}
_geo_cache: dict[str, dict | None] = {}
_geo_lock = asyncio.Lock()


async def _geocode(client: httpx.AsyncClient, lugar: str) -> dict | None:
    """Resuelve un nombre de lugar (cualquier parte del mundo) a coordenadas.
    Cachea resultados y respeta el limite de uso de Nominatim (1 req/s)."""
    key = (lugar or "").strip().lower()
    if not key:
        return None
    if key in _geo_cache:
        return _geo_cache[key]
    async with _geo_lock:
        if key in _geo_cache:
            return _geo_cache[key]
        # Fallback progresivo: intentamos la cadena completa (mas precisa) y, si
        # Nominatim no la ubica, vamos soltando la parte mas especifica de la
        # izquierda (p.ej. una direccion/POI) hasta que quede ciudad, depto, pais.
        partes = [p.strip() for p in lugar.split(",") if p.strip()]
        intentos = [", ".join(partes[i:]) for i in range(max(1, len(partes)))] or [lugar]
        res = None
        for intento in intentos:
            try:
                r = await client.get(_NOMINATIM, params={"q": intento, "format": "json", "limit": 1},
                                      headers=_UA, timeout=20)
                data = r.json() if r.status_code == 200 else []
            except Exception:
                data = []
            await asyncio.sleep(1.0)  # politica de uso Nominatim
            if data:
                top = data[0]
                res = {"nombre": top.get("display_name"), "lat": float(top["lat"]), "lon": float(top["lon"])}
                break
        _geo_cache[key] = res
        return res


async def _ruta(client: httpx.AsyncClient, o: dict, d: dict, con_geometria: bool) -> dict | None:
    ov = "full" if con_geometria else "false"
    url = f"{_OSRM}/{o['lon']},{o['lat']};{d['lon']},{d['lat']}"
    try:
        r = await client.get(url, params={"overview": ov, "geometries": "geojson"}, timeout=25)
        data = r.json()
    except Exception:
        return None
    if data.get("code") != "Ok" or not data.get("routes"):
        return None
    ruta = data["routes"][0]
    out = {"distancia_km": round(ruta["distance"] / 1000, 1), "duracion_min": round(ruta["duration"] / 60, 1)}
    if con_geometria:
        out["geometria"] = [[c[1], c[0]] for c in ruta["geometry"]["coordinates"]]  # [lat, lon] para Leaflet
    return out


@router.get("/geocode")
async def geocode_lugar(q: str, current_user: Usuario = Depends(get_current_user)):
    async with httpx.AsyncClient() as client:
        res = await _geocode(client, q)
    if not res:
        raise HTTPException(status_code=404, detail=f"No se encontró el lugar: {q}")
    return res


@router.get("/ruta")
async def calcular_ruta(
    origen: str,
    destino: str,
    current_user: Usuario = Depends(get_current_user),
):
    """Distancia y ruta por carretera entre dos lugares (consulta puntual)."""
    async with httpx.AsyncClient() as client:
        o = await _geocode(client, origen)
        if not o:
            raise HTTPException(404, f"Origen no encontrado: {origen}")
        d = await _geocode(client, destino)
        if not d:
            raise HTTPException(404, f"Destino no encontrado: {destino}")
        r = await _ruta(client, o, d, con_geometria=True)
    if not r:
        raise HTTPException(502, "No se pudo calcular la ruta entre esos puntos")
    return {"origen": o, "destino": d, **r}


@router.post("/ruta-masiva")
async def ruta_masiva(
    file: UploadFile = File(...),
    limite: int = Query(500, ge=1, le=2000),
    current_user: Usuario = Depends(get_current_user),
):
    """Carga un Excel con ORIGEN/DESTINO y devuelve otro con la distancia por
    carretera de cada par (Dijkstra sobre Contraction Hierarchies via OSRM)."""
    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo: {e}")
    df.columns = [str(c).strip() for c in df.columns]

    def _col(*cands):
        for c in cands:
            if c in df.columns:
                return c
        return None

    col_o = _col("ORIGEN", "Origen", "origen", "MUNICIPIO_ORIGEN", "CIUDAD_ORIGEN")
    col_d = _col("DESTINO", "Destino", "destino", "MUNICIPIO_DESTINO", "CIUDAD_DESTINO")
    if not col_o or not col_d:
        raise HTTPException(400, f"El archivo debe tener columnas ORIGEN y DESTINO. Columnas: {', '.join(df.columns)}")

    # Columnas opcionales para afinar la geocodificacion (direccion, depto/estado, pais)
    col_dir_o = _col("DIRECCION_ORIGEN", "DIR_ORIGEN", "DIRECCION ORIGEN")
    col_dir_d = _col("DIRECCION_DESTINO", "DIR_DESTINO", "DIRECCION DESTINO")
    col_dep_o = _col("DEPARTAMENTO_ORIGEN", "DEPTO_ORIGEN", "ESTADO_ORIGEN", "PROVINCIA_ORIGEN")
    col_dep_d = _col("DEPARTAMENTO_DESTINO", "DEPTO_DESTINO", "ESTADO_DESTINO", "PROVINCIA_DESTINO")
    col_pais_o = _col("PAIS_ORIGEN", "PAIS ORIGEN")
    col_pais_d = _col("PAIS_DESTINO", "PAIS DESTINO")

    def _componer(row, c_dir, c_mun, c_dep, c_pais) -> str:
        partes = []
        for c in (c_dir, c_mun, c_dep, c_pais):
            if not c:
                continue
            v = str(row[c]).strip()
            if v and v.lower() != "nan":
                partes.append(v)
        return ", ".join(partes)

    total_filas = len(df)
    df = df.head(limite).copy()

    dist_km: list = []
    dur_min: list = []
    estado: list = []
    async with httpx.AsyncClient() as client:
        for _, row in df.iterrows():
            o_txt = _componer(row, col_dir_o, col_o, col_dep_o, col_pais_o)
            d_txt = _componer(row, col_dir_d, col_d, col_dep_d, col_pais_d)
            if not o_txt or not d_txt:
                dist_km.append(None); dur_min.append(None); estado.append("FALTAN DATOS"); continue
            o = await _geocode(client, o_txt)
            d = await _geocode(client, d_txt)
            if not o:
                dist_km.append(None); dur_min.append(None); estado.append(f"ORIGEN no ubicado"); continue
            if not d:
                dist_km.append(None); dur_min.append(None); estado.append(f"DESTINO no ubicado"); continue
            r = await _ruta(client, o, d, con_geometria=False)
            if not r:
                dist_km.append(None); dur_min.append(None); estado.append("SIN RUTA"); continue
            dist_km.append(r["distancia_km"]); dur_min.append(r["duracion_min"]); estado.append("OK")

    df["DISTANCIA_KM"] = dist_km
    df["DURACION_MIN"] = dur_min
    df["ESTADO_RUTA"] = estado

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Distancias")
    output.seek(0)
    filename = f"TarifaX_distancias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ok = sum(1 for e in estado if e == "OK")
    return {
        "stats": {"filas": len(df), "total_archivo": total_filas, "calculadas": ok,
                  "sin_ruta": len(df) - ok, "truncado": total_filas > len(df)},
        "filename": filename,
        "file_base64": base64.b64encode(output.read()).decode("utf-8"),
    }


@router.get("/template")
async def descargar_plantilla(
    current_user: Usuario = Depends(get_current_user),
):
    if not TEMPLATE_PATH.exists():
        raise HTTPException(status_code=404, detail="Plantilla no encontrada en el servidor")
    with open(TEMPLATE_PATH, "rb") as f:
        content = f.read()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_cotizacion_tarifax.xlsx"},
    )


@router.get("/plantilla-distancias")
async def plantilla_distancias(current_user: Usuario = Depends(get_current_user)):
    """Genera un Excel con los encabezados esperados por el calculo masivo de
    distancias (obligatorios + opcionales) y dos filas de ejemplo."""
    cols = [
        "ORIGEN", "DEPARTAMENTO_ORIGEN", "PAIS_ORIGEN", "DIRECCION_ORIGEN",
        "DESTINO", "DEPARTAMENTO_DESTINO", "PAIS_DESTINO", "DIRECCION_DESTINO",
    ]
    ejemplo = pd.DataFrame([
        {"ORIGEN": "Bogotá", "DEPARTAMENTO_ORIGEN": "Cundinamarca", "PAIS_ORIGEN": "Colombia", "DIRECCION_ORIGEN": "",
         "DESTINO": "Medellín", "DEPARTAMENTO_DESTINO": "Antioquia", "PAIS_DESTINO": "Colombia", "DIRECCION_DESTINO": ""},
        {"ORIGEN": "Cali", "DEPARTAMENTO_ORIGEN": "Valle del Cauca", "PAIS_ORIGEN": "Colombia", "DIRECCION_ORIGEN": "",
         "DESTINO": "Barranquilla", "DEPARTAMENTO_DESTINO": "Atlántico", "PAIS_DESTINO": "Colombia", "DIRECCION_DESTINO": ""},
    ], columns=cols)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ejemplo.to_excel(writer, index=False, sheet_name="Distancias")
        ws = writer.sheets["Distancias"]
        for i, col in enumerate(cols):
            ws.column_dimensions[chr(65 + i)].width = max(16, len(col) + 2)
    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_distancias_tarifax.xlsx"},
    )


@router.post("/merge")
async def merge_tarifas(
    file: UploadFile = File(...),
    current_user: Usuario = Depends(get_current_user),
):
    content = await file.read()
    try:
        df2 = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {e}")

    df2.columns = [str(c).strip() for c in df2.columns]

    faltantes = [c for c in REQUIRED_DF2_KEYS if c not in df2.columns]
    if faltantes:
        available = ", ".join(df2.columns.tolist())
        raise HTTPException(
            status_code=400,
            detail=(
                f"Faltan columnas clave en el archivo: {', '.join(faltantes)}. "
                f"Se requieren {', '.join(REQUIRED_DF2_KEYS)} para cruzar por ruta y categoria de vehiculo. "
                f"Columnas disponibles: {available}"
            ),
        )

    df1 = _load_df1()

    # Traduce la tipologia interna de la empresa (p.ej. TRACTOCAMION) al codigo
    # SICETAC (p.ej. 3S2) segun el mapeo configurado, para que el cruce y el CPK calcen.
    mapeo = _mapeo_norm()
    vehiculos_mapeados = 0
    if mapeo and "TIPO_VEHICULO" in df2.columns:
        tvn = _norm(df2["TIPO_VEHICULO"])
        nuevos = []
        for i in range(len(df2)):
            dest = mapeo.get(tvn.iloc[i])
            if dest:
                nuevos.append(dest); vehiculos_mapeados += 1
            else:
                nuevos.append(df2["TIPO_VEHICULO"].iloc[i])
        df2["TIPO_VEHICULO"] = nuevos

    # Columnas originales del archivo del cliente (para la hoja de no-coincidencias).
    df2_cols = list(df2.columns)

    # Llaves efectivas: solo las que existen en ambos archivos...
    keys = [(a, b) for (a, b) in JOIN_KEYS if a in df2.columns and b in df1.columns]
    # ...y descartar las que vienen totalmente en blanco en el archivo del cliente
    # (p.ej. CARROCERIA vacia), para no forzar todo a "sin coincidencia".
    keys = [(a, b) for (a, b) in keys if _norm(df2[a]).ne("").any()]

    # Columnas normalizadas de cruce en el archivo del cliente.
    norm_cols: list[str] = []
    for i, (a, _) in enumerate(keys):
        nc = f"__k{i}"
        df2[nc] = _norm(df2[a])
        norm_cols.append(nc)

    grouped = _grouped_df1(keys)

    # Cruce 1:1 por la ruta + categoria de vehiculo consultada.
    result = pd.merge(df2, grouped, on=norm_cols, how="left", suffixes=("", "_sicetac"))
    result = result.drop(columns=norm_cols)
    result["procesado_en"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if COL_PRECIO_ACTUAL in result.columns and COL_PRECIO_SICETAC in result.columns:
        precio_cli = pd.to_numeric(result[COL_PRECIO_ACTUAL], errors="coerce")
        precio_sic = pd.to_numeric(result[COL_PRECIO_SICETAC], errors="coerce").replace(0, pd.NA)
        result["variacion_precio"] = (precio_cli / precio_sic).round(4)

    total = len(result)
    if COL_PRECIO_SICETAC in result.columns:
        matched_mask = result[COL_PRECIO_SICETAC].notna()
    else:
        matched_mask = pd.Series([False] * total, index=result.index)
    cruzados = int(matched_mask.sum())
    unmatched = total - cruzados
    match_rate = round(cruzados / total * 100, 1) if total > 0 else 0.0

    # ── Hoja 2 "calculo por cpk": rutas SIN coincidencia con tarifa teorica ──
    # Tarifa teorica = distancia(origen, destino) x CPK promedio del municipio origen.
    cpk_map, pivot_cpk = _cpk_por_origen()
    dist_map = _distancia_origen_destino()

    sin = result[~matched_mask].copy()
    cpk_cols = [c for c in df2_cols if c in sin.columns]
    cpk_sheet = sin[cpk_cols].reset_index(drop=True)
    n = len(cpk_sheet)

    o_norm = _norm(cpk_sheet["ORIGEN"]) if "ORIGEN" in cpk_sheet.columns else pd.Series([""] * n)
    d_norm = _norm(cpk_sheet["DESTINO"]) if "DESTINO" in cpk_sheet.columns else pd.Series([""] * n)
    v_norm = _norm(cpk_sheet["TIPO_VEHICULO"]) if "TIPO_VEHICULO" in cpk_sheet.columns else pd.Series([""] * n)

    # Distancia por carretera con el MISMO motor del modulo de Distancias
    # (geocodificacion + OSRM / Dijkstra sobre Contraction Hierarchies) para las
    # rutas sin coincidencia. Es mas exacta que el promedio de SICETAC al
    # multiplicar CPK x distancia. Se calcula una sola vez por ruta unica y se
    # apoya en el cache de geocodificacion (los municipios se repiten mucho).
    LIMITE_RUTEO = 400
    routing_km: dict = {}
    if "ORIGEN" in cpk_sheet.columns and "DESTINO" in cpk_sheet.columns:
        rutas_unicas: dict = {}
        for i in range(n):
            k = (o_norm.iloc[i], d_norm.iloc[i])
            if not o_norm.iloc[i] or not d_norm.iloc[i] or k in rutas_unicas:
                continue
            rutas_unicas[k] = (str(cpk_sheet["ORIGEN"].iloc[i]).strip(),
                               str(cpk_sheet["DESTINO"].iloc[i]).strip())

        def _con_pais(t: str) -> str:
            # Sesga la geocodificacion a Colombia (carga domestica) si no se
            # especifico ya un pais/detalle con coma.
            return t if "," in t else f"{t}, Colombia"

        async with httpx.AsyncClient() as client:
            for k, (o_txt, d_txt) in list(rutas_unicas.items())[:LIMITE_RUTEO]:
                if not o_txt or not d_txt or o_txt.lower() == "nan" or d_txt.lower() == "nan":
                    continue
                o = await _geocode(client, _con_pais(o_txt))
                d = await _geocode(client, _con_pais(d_txt))
                if not o or not d:
                    continue
                r = await _ruta(client, o, d, con_geometria=False)
                if r:
                    routing_km[k] = r["distancia_km"]

    # Distancia por fila. Prioridad: motor de rutas (mas exacto) > archivo cliente > SICETAC.
    if COL_DISTANCIA in cpk_sheet.columns:
        dist_cli = pd.to_numeric(cpk_sheet[COL_DISTANCIA], errors="coerce").astype("Float64")
    else:
        dist_cli = pd.Series([pd.NA] * n, dtype="Float64")
    dist_sicetac = pd.Series(
        [dist_map.get((o_norm.iloc[i], d_norm.iloc[i])) for i in range(n)], dtype="Float64"
    )
    dist_ruta = pd.Series(
        [routing_km.get((o_norm.iloc[i], d_norm.iloc[i])) for i in range(n)], dtype="Float64"
    )
    dist_final = dist_ruta.fillna(dist_cli).fillna(dist_sicetac)

    def _fuente(i: int) -> str:
        if pd.notna(dist_ruta.iloc[i]): return "Ruta por carretera (OSRM)"
        if pd.notna(dist_cli.iloc[i]): return "Archivo cliente"
        if pd.notna(dist_sicetac.iloc[i]): return "Promedio SICETAC"
        return "Sin dato"

    # CPK segun (municipio origen, tipologia de vehiculo) de cada fila.
    cpk_series = pd.Series(
        [cpk_map.get((o_norm.iloc[i], v_norm.iloc[i])) for i in range(n)], dtype="Float64"
    )

    cpk_sheet["DISTANCIA_KM"] = dist_final.round(1)
    cpk_sheet["FUENTE_DISTANCIA"] = [_fuente(i) for i in range(n)]
    cpk_sheet["CPK_PROMEDIO_ORIGEN"] = cpk_series.round(2)
    cpk_sheet["TARIFA_TEORICA_CPK"] = (dist_final * cpk_series).round(0)
    con_teorica = int(cpk_sheet["TARIFA_TEORICA_CPK"].notna().sum())
    dist_ruteadas = int(dist_ruta.notna().sum())

    matched = result[matched_mask]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        matched.to_excel(writer, index=False, sheet_name="TarifaX_Resultado")
        cpk_sheet.to_excel(writer, index=False, sheet_name="calculo por cpk")
        pivot_cpk.to_excel(writer, index=False, sheet_name="CPK por Origen")
    output.seek(0)

    filename = f"TarifaX_resultado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return {
        "stats": {
            "registros": total,
            "cruzados": cruzados,
            "sin_coincidencia": unmatched,
            "tasa_cruce": match_rate,
            "llaves_cruce": [a for a, _ in keys],
            "tarifa_teorica_calculada": con_teorica,
            "municipios_origen_cpk": len(pivot_cpk),
            "vehiculos_mapeados": vehiculos_mapeados,
            "distancias_por_ruta": dist_ruteadas,
        },
        "preview": {
            "cruzados": _preview(matched),
            "calculo_por_cpk": _preview(cpk_sheet),
        },
        "filename": filename,
        "file_base64": base64.b64encode(output.read()).decode("utf-8"),
    }
